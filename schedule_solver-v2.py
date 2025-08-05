# --------------------------------------------------------------------------
# 排班系統 v2.7 (最終註解版) 補上同儕公平機制
# --------------------------------------------------------------------------
# 專案說明：
# 本系統旨在解決一個複雜的醫師排班問題，
# 綜合考量多種硬性規定與人性化的軟性目標，
# 以找出最理想的排班解方。
#
# 核心功能：
# 1. 遵守所有硬性規則 (Hard Constraints)。
# 2. 透過多層次的軟性目標 (Soft Constraints) 找出最佳解：
#    - 主要目標：最大化總使用點數、最大化排班分散度、強力懲罰過短間隔。
#    - 次要目標：在點數無法用完時，追求同儕間的公平性。
#    - 最終目標：以極低權重獎勵醫師在自己區域工作，作為打破平局的依據。
# 3. 自動生成一份精美、多頁籤、帶有顏色標記的視覺化 Excel 報告。

# 輸出範例： 目前最好的解 (12月 2025) 再來就是10月第二好 再來是5月 (其中只有 2025 2月 6月 7月 9月 無法把點數用完)
#   - 總使用點數       :   143 (分數: 1430000) // 目前測資 總點數最大就是143
#   - 線性間隔獎勵      :  4900 (分數: 49000) // 這也是目前最大的4900
#   - 隔兩天次數(懲罰)   :     0 (分數: 0) // 0是最好的
#   - 同儕公平性(懲罰)   :     0 (分數: 0) // 0是最好的
#   - 總排班數量       :   112 (分數: 11000) // 112 是目前最大的 (應該可以更大)
#   - I 區優先獎勵     :    19 (分數: 190) //19也是目前看到最大的
#   - 在家區域獎勵      :   109 (分數: 10) //這可以更高 但是其他的 已經算是最頂了

# --------------------------------------------------------------------------

# 步驟 1: 匯入所有必要的工具程式庫
import pandas as pd
from ortools.sat.python import cp_model
import holidays # 用於自動抓取國定假日
import os
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict # 用於建立帶有預設值的字典，方便分組

# --------------------------------------------------------------------------
# Section A: 輔助工具定義
# --------------------------------------------------------------------------

class SolutionCounter(cp_model.CpSolverSolutionCallback):
    """
    這是一個繼承自 OR-Tools 內建功能的輔助類別。
    它的主要作用是：每當求解器找到一個「可行解」(滿足所有硬性規定的解)時，
    就會自動觸發 on_solution_callback 函式，讓我們可以即時印出該解的詳細資訊。
    """
    def __init__(self, objectives, weights):
        super().__init__()
        self._solution_count = 0
        self._objectives = objectives
        self._weights = weights
        # 預先定義好各個目標的輸出順序與顯示名稱，讓終端機的輸出整齊一致。
        self._display_order = [
            ('total_used_points', '總使用點數'),
            ('linear_gaps_bonus', '線性間隔獎勵'),
            ('min_gap_count', '隔兩天次數(懲罰)'),
            ('fairness_penalty', '同儕公平性(懲罰)'),
            ('total_shifts_filled', '總排班數量'),
            ('i_priority_bonus', 'I 區優先獎勵'),
            ('home_area_bonus', '在家區域獎勵')
        ]

    def on_solution_callback(self):
        """每找到一個可行解，此函式就會被呼叫一次。"""
        self._solution_count += 1
        print(f"\n--- 找到第 {self._solution_count} 個可行解 ---")

        total_score = 0
        
        # 使用一個乾淨的迴圈，依據預設的順序，處理並印出每個目標的分數。
        for key, display_name in self._display_order:
            if key in self._objectives:
                # 從求解器中獲取此解中該目標的原始數值 (例如：總共排了幾天班)
                raw_val = self.Value(self._objectives[key])
                # 計算加權後的分數
                score = raw_val * self._weights[key]
                # 累加到總分
                total_score += score
                # 使用f-string格式化輸出，讓版面整齊，並用int()確保分數顯示為整數。
                print(f"  - {display_name:<12}: {raw_val:>5} (分數: {int(score)})")
        
        print(f"  >> 此解總分: {int(total_score)}")

    def solution_count(self):
        """提供一個方法來獲取目前找到的總可行解數量。"""
        return self._solution_count

def format_excel(writer, doctor_schedule_df, weekend_days, official_holidays, doctor_info):
    """
    這個函式負責將最終的排班結果美化成一份專業的 Excel 報告。
    它利用 openpyxl 工具庫來對 Excel 的儲存格進行詳細的格式設定。
    """
    # 獲取 Excel writer 中的主要工作簿(workbook)和工作表(worksheet)物件
    workbook = writer.book
    ws_doctor = writer.sheets['醫師月曆班表']

    # 定義各區域的代表顏色，並建立對應的填滿樣式
    colors = {'A': 'ADD8E6', 'B': '90EE90', 'C': 'FFFFE0', 'I': 'FFB6C1'}
    fills = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in colors.items()}

    # 定義週末、國定假日、預休日、表頭的特殊填滿樣式
    weekend_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type="solid") # 淺灰色
    holiday_fill = PatternFill(start_color='FFDDC1', end_color='FFDDC1', fill_type="solid") # 淺橘色
    unavailable_fill = PatternFill(patternType='gray0625', fgColor='A9A9A9') # 深灰色網底
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type="solid") # 淺藍色

    # 定義通用的字體、對齊和框線樣式
    header_font = Font(bold=True, color='000000')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 為週末和國定假日的整欄上色
    for col_idx, day in enumerate(doctor_schedule_df.columns, 2):
        fill_to_apply = None
        if day in official_holidays: fill_to_apply = holiday_fill
        elif day in weekend_days: fill_to_apply = weekend_fill
        if fill_to_apply:
            for row_idx in range(1, ws_doctor.max_row + 1):
                ws_doctor.cell(row=row_idx, column=col_idx).fill = fill_to_apply

    # 遍歷所有資料儲存格，套用對應的樣式
    for r_idx, doc in enumerate(doctor_schedule_df.index, 2):
        ws_doctor.cell(row=r_idx, column=1).font = Font(bold=True) # 醫師姓名加粗
        ws_doctor.row_dimensions[r_idx].height = 25 # 設定行高
        for c_idx, day in enumerate(doctor_schedule_df.columns, 2):
            cell = ws_doctor.cell(row=r_idx, column=c_idx)
            cell.alignment = center_align
            cell.border = thin_border
            # 如果儲存格有值(代表有排班)，則根據區域填上對應顏色
            if cell.value and cell.value in fills:
                cell.fill = fills[cell.value]

    # 標記醫師的預休日
    doc_to_row_map = {doc: i + 2 for i, doc in enumerate(doctor_schedule_df.index)}
    for doc, info in doctor_info.items():
        row_idx = doc_to_row_map[doc]
        for day_off in info['不可排班日']:
            if day_off in doctor_schedule_df.columns:
                cell = ws_doctor.cell(row=row_idx, column=day_off + 1)
                cell.fill = unavailable_fill
                cell.value = "預休"
                cell.font = Font(color='FFFFFF', bold=True)

    # 格式化表頭
    for col_idx in range(1, ws_doctor.max_column + 1):
        cell = ws_doctor.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
        ws_doctor.column_dimensions[get_column_letter(col_idx)].width = 6 # 設定欄寬
    ws_doctor.row_dimensions[1].height = 20
    ws_doctor.column_dimensions['A'].width = 12 # 醫師姓名的欄位寬一點

    # 美化「點數統計總覽」工作表
    ws_summary = writer.sheets['點數統計總覽']
    for col_idx in range(1, ws_summary.max_column + 1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 15
        ws_summary.cell(row=1, column=col_idx).font = header_font
        ws_summary.cell(row=1, column=col_idx).fill = header_fill

# --------------------------------------------------------------------------
# Section B: 主要排班函式
# --------------------------------------------------------------------------

def solve_schedule():
    """
    這是整個排班系統的主體函式。
    它會依序執行：資料準備 -> 建立模型 -> 設定規則 -> 求解 -> 輸出報告。
    """
    # --- 步驟 1: 資料準備與環境設定 ---
    # 將醫師資料直接寫在程式碼中，方便攜帶與執行。
    csv_data = """
醫師姓名,區域,點數上限,不可排班日
如,A,8,"26,27"
秀,A,8,"1,2,5,6"
橋,A,6,"1,2,3,4,5,6,7,8,9,19,20"
君,A,6,"4"
翔,A,6,"1,3,4"
航,A,8,"1,14,15,16,17,18,19,20"
淇,B,8,"1,2,25,28"
慈,B,8,"3,4"
恩,B,8,""
屹,B,8,"4,5"
軒,B,6,"2,3,5"
佑,C,8,""
翰,C,6,"1,2,3,4,5,6,7,8,9,13,27"
潔,C,5,"16,17,18,19,20,21,22,23,24,25,26,27,28,29,30"
諺,C,5,"1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,26"
宣,C,8,"26,27"
韶,C,8,"2,3,4,5,6,7,8"
然,I,8,"1,2,3,4,5,6"
偉,I,8,"1,2,4,5"
煒,I,7,"21,22,23,24,25,26,27,28,29,30"
"""
    print("1. 正在讀取內建的醫師資料...")
    # 使用 io.StringIO 將字串偽裝成檔案，讓 pandas 可以讀取
    data_io = io.StringIO(csv_data)
    df = pd.read_csv(data_io, engine='python')
    df.columns = df.columns.str.strip() # 去除欄位名稱前後的空白
    # 將「不可排班日」字串轉換為數字列表，並處理空值
    df['不可排班日'] = df['不可排班日'].fillna('').astype(str).apply(
        lambda x: [int(d) for d in x.split(',') if d.strip().isdigit()]
    )
    # 設定要排班的年份和月份
    YEAR, MONTH = 2025, 10
    num_days = pd.Period(f'{YEAR}-{MONTH}-01').days_in_month
    # 找出該月份中所有的週末(週六、週日)和國定假日
    date_range = pd.to_datetime(pd.Series(pd.date_range(start=f'{YEAR}-{MONTH}-01', end=f'{YEAR}-{MONTH}-{num_days}')))
    weekend_days = date_range[date_range.dt.dayofweek.isin([5, 6])].dt.day.tolist()
    official_holidays = [d.day for d in holidays.TW(years=YEAR) if d.month == MONTH]
    # 將週末和國定假日合併為「雙倍點數日」
    double_point_days = set(weekend_days + official_holidays)
    doctors = df['醫師姓名'].tolist()
    areas = ['A', 'B', 'C', 'I']
    # 將 DataFrame 轉換為字典，方便後續快速查找醫師資訊
    doctor_info = df.set_index('醫師姓名').to_dict('index')

    # --- 步驟 2: 建立數學模型與決策變數 ---
    print("2. 正在建立數學模型...")
    model = cp_model.CpModel()
    
    # 建立核心決策變數 `shifts`
    # 這是一個巨大的字典，key是 (醫師, 日期, 區域) 的組合
    # value是 OR-Tools 的布林變數，求解器的工作就是決定這些變數是 0 (不排班) 還是 1 (排班)
    shifts = {}
    for doc in doctors:
        for day in range(1, num_days + 1):
            for area in areas:
                shifts[(doc, day, area)] = model.NewBoolVar(f'shift_{doc}_{day}_{area}')

    # --- 步驟 3: 加入所有「必須遵守」的硬性規則 ---
    # 硬性規則是班表成立的底線，任何解都不能違反。

    # 規則：一個班次(特定日期、特定區域)最多只能有一位醫師。
    for day in range(1, num_days + 1):
        for area in areas: model.AddAtMostOne([shifts[(doc, day, area)] for doc in doctors])
    
    # 規則：一位醫師一天最多只能上一個班(不能同時在A區和B區上班)。
    for day in range(1, num_days + 1):
        for doc in doctors: model.AddAtMostOne([shifts[(doc, day, area)] for area in areas])
    
    # 規則：只有I區的醫師才能在I區排班。
    i_doctors = [d for d, info in doctor_info.items() if info['區域'] == 'I']
    support_doctors = [d for d, info in doctor_info.items() if info['區域'] in ['A', 'B', 'C']]
    for day in range(1, num_days + 1): model.Add(sum(shifts[(doc, day, 'I')] for doc in support_doctors) == 0)
    
    # 規則：醫師排班後，接下來的兩天必須休息。
    # 檢查一個連續三天的滑動窗口，確保其中的班次總數 <= 1。
    for doc in doctors:
        for day in range(1, num_days - 1): model.Add(sum(shifts[doc, d, area] for area in areas for d in range(day, day+3)) <= 1)
    
    # 規則：不可將醫師排在他預休的日子。
    for doc in doctors:
        for day in doctor_info[doc]['不可排班日']:
            if 1 <= day <= num_days:
                for area in areas: model.Add(shifts[(doc, day, area)] == 0)

    # 規則：每位醫師的總點數(平日1點,假日2點)不可超過其上限。
    points_per_doctor = {}
    for doc in doctors:
        points_for_doc = sum(shifts[(doc, day, area)] * (2 if day in double_point_days else 1) for day in range(1, num_days + 1) for area in areas)
        points_per_doctor[doc] = points_for_doc
        model.Add(points_per_doctor[doc] <= doctor_info[doc]['點數上限'])

    # --- 步驟 4: 設定系統的「優化目標」(軟性條件) ---
    # 軟性目標是用來評價一個「可行解」好壞的標準，系統會盡力找到讓總分最高的解。
    
    # 輔助變數：`is_work_day` 方便後續的目標計算。
    is_work_day = {}
    for doc in doctors:
        for day in range(1, num_days + 1):
            is_work_day[doc, day] = model.NewBoolVar(f'is_work_day_{doc}_{day}')
            model.Add(is_work_day[doc, day] == sum(shifts[doc, day, area] for area in areas))

    # 目標1: 總使用點數。希望系統盡可能把所有人的點數用完，讓班表飽滿。
    total_used_points = sum(points_per_doctor.values())
    
    # 目標2: 線性間隔獎勵。計算每兩次連續排班的間隔天數d，給予 10*d 的獎勵。
    # 這是通用性的分散度獎勵，鼓勵班表拉得越開越好。
    all_linear_bonuses = []
    for doc in doctors:
        for d1 in range(1, num_days + 1):
            for d2 in range(d1 + 1, num_days + 1):
                # 判斷 d1 和 d2 是否為「連續」的排班日 (中間沒有其他班)
                is_consecutive = model.NewBoolVar(f'consecutive_{doc}_{d1}_{d2}')
                no_work_in_between_literals = [is_work_day[doc, d].Not() for d in range(d1 + 1, d2)]
                model.AddBoolAnd([is_work_day[doc, d1], is_work_day[doc, d2]] + no_work_in_between_literals).OnlyEnforceIf(is_consecutive)
                model.AddBoolOr([is_work_day[doc, d1].Not(), is_work_day[doc, d2].Not()] + [is_work_day[doc, d] for d in range(d1 + 1, d2)]).OnlyEnforceIf(is_consecutive.Not())
                gap = d2 - d1
                linear_bonus = 10 * gap 
                all_linear_bonuses.append(is_consecutive * linear_bonus)
    total_linear_gaps_bonus = sum(all_linear_bonuses)

    # 目標3: 精準懲罰「隔兩天」的排班。計算「班-休-休-班」模式的發生次數。
    # 這是為了強力避免出現剛好滿足最低休息天數的班表。
    min_gap_penalties = []
    for doc in doctors:
        for day in range(1, num_days - 2):
            has_min_gap = model.NewBoolVar(f'has_min_gap_{doc}_{day}')
            model.AddBoolAnd([is_work_day[doc, day], is_work_day[doc, day + 3]]).OnlyEnforceIf(has_min_gap)
            model.AddBoolOr([is_work_day[doc, day].Not(), is_work_day[doc, day + 3].Not()]).OnlyEnforceIf(has_min_gap.Not())
            min_gap_penalties.append(has_min_gap)
    total_min_gap_count = sum(min_gap_penalties)

    # 目標4: 同儕公平性。懲罰同一個群組內(同區、同點數上限)醫師們最終點數的差距。
    # 這是為了在點數無法用完時，讓減班的狀況盡可能公平地分攤。
    peer_groups = defaultdict(list)
    for doc, info in doctor_info.items():
        key = (info['區域'], info['點數上限'])
        peer_groups[key].append(doc)
    all_ranges = []
    for group_key, group_docs in peer_groups.items():
        if len(group_docs) > 1:
            group_points = [points_per_doctor[doc] for doc in group_docs]
            min_points = model.NewIntVar(0, 100, f'min_points_{group_key}')
            max_points = model.NewIntVar(0, 100, f'max_points_{group_key}')
            model.AddMinEquality(min_points, group_points)
            model.AddMaxEquality(max_points, group_points)
            group_range = model.NewIntVar(0, 100, f'range_{group_key}')
            model.Add(group_range == max_points - min_points)
            all_ranges.append(group_range)
    fairness_penalty = sum(all_ranges)

    # 其他次要目標
    total_shifts_filled = sum(shifts.values())
    i_priority_bonus = sum(shifts[(doc, day, 'I')] for doc in i_doctors for day in range(1, num_days + 1))
    home_area_bonus = sum(shifts[(doc, day, info['區域'])] for doc, info in doctor_info.items() for day in range(1, num_days + 1))

    # --- 步驟 4.1: 整合所有目標與其權重 ---
    # `objectives` 字典儲存了所有目標的計算結果變數。
    objectives = {
        'total_used_points': total_used_points, # 總使用點數
        'linear_gaps_bonus': total_linear_gaps_bonus, # 線性間隔獎勵 (相隔d天 總分加 10*d) 已經有線性加乘 所以下面的權重可以小一點
        'min_gap_count': total_min_gap_count, # 隔兩天次數(懲罰) (班-休-休-班 的次數)
        'fairness_penalty': fairness_penalty, # 同儕公平性懲罰 (同一群組內醫師點數 最大和最小 差距 總和)
        'total_shifts_filled': total_shifts_filled, # 總排班數量 (所有醫師的班次總和)
        'i_priority_bonus': i_priority_bonus, # I 區優先獎勵 (I區醫師的班次總和)
        'home_area_bonus': home_area_bonus, # 在家區域獎勵 (所有醫師在自己區域的班次總和)
    }
    
    # `weights` 字典定義了每個目標的重要性，這是整個系統決策的核心。
    # 正數代表獎勵(越大越好)，負數代表懲罰(越接近0越好)。
    # 權重的絕對值大小決定了優先級。
    weights = {
        'total_used_points': 10000,   # 最高優先級：讓大家有班上、點數用得多。
        'linear_gaps_bonus': 10,      # 次要獎勵：普遍性地鼓勵班表分散。 因為 上面的 total_linear_gaps_bonus 已經 有線性加成了 所以這邊的權重可以小一點
        'min_gap_count': -500,        # 強力懲罰：非常不希望看到隔兩天的班。
        'fairness_penalty': -200,     # 次要懲罰：如果不滿班，希望大家公平地減班。
        'total_shifts_filled': 100,   # 一般獎勵：在滿足前述條件下，盡量填滿所有空班。
        'i_priority_bonus': 10,       # 通用獎勵：讓I區醫師優先在I區上班。
        'home_area_bonus': 0.1,       # 最低權重獎勵：所有條件都差不多時，才考慮讓醫師在自己區上班。
    }

    # 設定模型的總目標：最大化 (所有目標 * 對應權重) 的總和。
    model.Maximize(sum(objectives[name] * weights[name] for name in objectives))

    # --- 步驟 5: 啟動求解器 ---
    solver = cp_model.CpSolver()
    
    # 設定求解器參數：回報所有找到的可行解，而不只是找到最優解後就停止。
    solver.parameters.enumerate_all_solutions = True
    # 建立我們自訂的 SolutionCounter，將目標和權重傳入，以便在找到解時印出詳細報告。
    solution_counter = SolutionCounter(objectives, weights)
    
    # 設定最長運算時間，避免無窮盡的搜索。
    solver.parameters.max_time_in_seconds = 120.0
    print(f"3. 正在運算，尋找所有可能的排班方案 ({YEAR}-{MONTH})...")
    # 執行求解！
    status = solver.Solve(model, solution_counter)

    # --- 步驟 6: 處理結果並生成報告 ---
    # 檢查求解器的最終狀態
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        print("4. 已找到最佳解！正在生成視覺化 Excel 報告...")

        # 建立空的 DataFrame 來存放最終班表結果
        schedule_df = pd.DataFrame({area: [""] * num_days for area in areas}, index=range(1, num_days + 1))
        doctor_schedule_df = pd.DataFrame('', index=doctors, columns=range(1, num_days + 1))
        points_summary_data = []

        # 遍歷所有決策變數，找出值為1的(代表有排班)，並填入 DataFrame
        for doc in doctors:
            points_used, days_worked = 0, []
            for day in range(1, num_days + 1):
                for area in areas:
                    if solver.Value(shifts[(doc, day, area)]) == 1:
                        schedule_df.loc[day, area] = doc
                        doctor_schedule_df.loc[doc, day] = area
                        points_used += 2 if day in double_point_days else 1
                        days_worked.append(f"{day}({area})")
            # 彙總每位醫師的統計資料
            points_summary_data.append({
                '醫師姓名': doc, '區域': doctor_info[doc]['區域'], '點數上限': doctor_info[doc]['點數上限'],
                '實際點數': points_used, '剩餘點數': doctor_info[doc]['點數上限'] - points_used,
                '排班日與區域': ", ".join(days_worked)
            })
        points_summary_df = pd.DataFrame(points_summary_data)

        # 使用 pd.ExcelWriter 將多個 DataFrame 寫入到同一個 Excel 檔案的不同工作表
        output_filename = f'schedule_result_{YEAR}-{MONTH}.xlsx'
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            doctor_schedule_df.to_excel(writer, sheet_name='醫師月曆班表')
            points_summary_df.to_excel(writer, sheet_name='點數統計總覽', index=False)
            schedule_df.to_excel(writer, sheet_name='區域班表')
            # 呼叫我們定義的函式來美化 Excel 格式
            format_excel(writer, doctor_schedule_df, weekend_days, official_holidays, doctor_info)

        print(f"\n✅ **排班完成！**")
        print(f"   詳細結果請見您資料夾中的 **{output_filename}** 檔案。")
        print("\n--- 最終排班結果分析 ---")
        print(f"在所有規則限制下，系統總共找到了 **{solution_counter.solution_count()}** 種不同的可行排班方案。")
        print(f"呈現的是其中一個綜合評分最高的「最佳解」。")

    else:
        print(f"\n❌ **錯誤：** 在目前的規則下，找不到任何可行的排班解。")

# --------------------------------------------------------------------------
# Section C: 程式執行入口
# --------------------------------------------------------------------------
# 確保這個腳本是直接被執行，而不是被其他檔案匯入時，才啟動排班。
if __name__ == '__main__':
    solve_schedule()